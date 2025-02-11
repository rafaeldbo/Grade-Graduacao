import pandas as pd
import openpyxl

from os import path

from .config_grade import load_configs
from .data_processing import load_space_data
from .utils import warning, info, error, success, time_to_integer, get_digit, time_differece, float_to_time
from .settings import *

  
def construct_calendar(abs_path:str) -> None:
    
    CONFIGS = load_configs(abs_path)
    DATA = load_space_data(abs_path, CONFIGS)
    
    timetable_height = time_to_integer(CONFIGS['FIXED_END_HOURS'][-1]) + 6
    timetable_width = 22
    
    def conflit(sheet:Worksheet, start_col: int, start_row: int, end_col: int, end_row: int) -> bool:
        for row in range(start_row, end_row+1):
            for col in range(start_col, end_col+1):
                cell = sheet.cell(row=row, column=col)
                if (cell.value is not None) or cell.fill.start_color.rgb != WHITE:
                    return False
        return True

    def slot_class(sheet:Worksheet, data:dict, timetable_start_col: int, timetable_start_row: int, show_professors: bool=False, show_errors: bool=False) -> None:
        start_col = col_letter(timetable_start_col + DAY_PARSER[data['dia_semana']]*4 + (4 if (data['posicao'] == 2) else 2))
        end_col = col_letter(timetable_start_col + DAY_PARSER[data['dia_semana']]*4 + (2 if (data['posicao'] == 1) else 4)) 
        start_row = time_to_integer(data['hora_inicio']) + timetable_start_row + 6
        end_row = time_to_integer(data['hora_fim']) + timetable_start_row + 5
        
        if not conflit(sheet, col_number(start_col), start_row, col_number(end_col), end_row):
            if show_errors:
                error(f'Não foi possível adicionar a {data["tipo_atividade"]} da disciplina {data["nome_disciplina"]} da turma {data["curso"]}-{get_digit(data["serie"])}{data["turma"]} no horário {data["hora_inicio"]} às {data["hora_fim"]} da {data["dia_semana"]} devido a um conflito com outro disciplina já adicionada')
            return
        duration = time_differece(data['hora_inicio'], data['hora_fim'])
        if not (duration == 2 or (duration == 1.5 and data['nome_disciplina'] in CONFIGS['DISCIPLINES_4_SLOTS_CLASS']) or (duration > 3)):
            if show_errors:
                error(f'Não foi possível adicionar a {data["tipo_atividade"]} da disciplina {data["nome_disciplina"]} da turma {data["curso"]}-{get_digit(data["serie"])}{data["turma"]} pois a duração da aula é diferente de 2h')
                print(f'Horário: {data["hora_inicio"]} às {data["hora_fim"]} da {data["dia_semana"]} | Duração: {float_to_time(duration)}')
            return
        
        color = CLASS_COLORS[data['cor']] if data['cor'] < 6 else CLASS_COLORS[-1]
        
        slot_title_cell = merge_cells(sheet, f'{start_col}{start_row}:{end_col}{start_row+2}')
        cell_styles(slot_title_cell, data['nome_disciplina'], FONT_BOLD_GRAY7, color, TOP_ALIGNMENT)
        
        slot_middle_cell = merge_cells(sheet, f'{start_col}{start_row+3}:{end_col}{start_row+3}')
        if data['nome_disciplina'] in CONFIGS['DISCIPLINES_4_SLOTS_CLASS']:
            cell_styles(slot_middle_cell, f'{data["hora_inicio"]} às {data["hora_fim"]}', FONT_BOLD_RED7)
        slot_middle_cell.fill = color
                
        slot_footer_cell = merge_cells(sheet, f'{start_col}{start_row+4}:{end_col}{end_row}')
        cell_styles(slot_footer_cell, font=FONT_BASE7, fill=color, alignment=BOTTOM_ALIGNMENT)
        if show_professors or (data['nome_disciplina'] in CONFIGS['EXCLUSIVE_TIMETABLE']): 
            slot_footer_cell.value = data["titular"]

    def slot_attendance(sheet:Worksheet, data:dict, timetable_start_col: int, timetable_start_row: int) -> None:
        start_col = col_letter(timetable_start_col + DAY_PARSER[data['dia_semana']]*4 + (4 if (data['posicao'] == 2) else 2))
        end_col = col_letter(timetable_start_col + DAY_PARSER[data['dia_semana']]*4 + (2 if (data['posicao'] == 1) else 4)) 
        start_row = time_to_integer(data['hora_inicio']) + timetable_start_row + 6
        end_row = time_to_integer(data['hora_fim']) + timetable_start_row + 5
        
        if not conflit(sheet, col_number(start_col), start_row, col_number(end_col), end_row):
            error(f'Não foi possível adicionar a {data["tipo_atividade"]} da disciplina {data["nome_disciplina"]} da turma {data["curso"]}-{get_digit(data["serie"])}{data["turma"]} no horário {data["hora_inicio"]} às {data["hora_fim"]} da {data["dia_semana"]} devido a um conflito com outro disciplina já adicionada')
            return
        duration = time_differece(data['hora_inicio'], data['hora_fim'])
        if duration != 1.5:
            error(f'Não foi possível adicionar a {data["tipo_atividade"]} da disciplina {data["nome_disciplina"]} da turma {data["curso"]}-{get_digit(data["serie"])}{data["turma"]} pois a duração da aula é diferente de 1h30')
            print(f'Horário: {data["hora_inicio"]} às {data["hora_fim"]} da {data["dia_semana"]} | Duração: {float_to_time(duration)}')
        
        color = ATTENDANCE_BORDERS[data['cor']] if data['cor'] < 6 else ATTENDANCE_BORDERS[-1]
        
        slot_title_cell = merge_cells(sheet, f'{start_col}{start_row}:{end_col}{start_row+1}')
        cell_styles(slot_title_cell, 'Horário de Atendimento', FONT_BOLD_GRAY7, alignment=TOP_ALIGNMENT)
        
        slot_middle_cell = merge_cells(sheet, f'{start_col}{start_row+2}:{end_col}{start_row+3}')
        cell_styles(slot_middle_cell, data['nome_disciplina'], FONT_BASE7, alignment=CENTER_ALIGNMENT)
        
        slot_footer_cell = merge_cells(sheet, f'{start_col}{start_row+4}:{end_col}{end_row}')
        cell_styles(slot_footer_cell,  f'{data["hora_inicio"]} às {data["hora_fim"]}', FONT_BASE7, alignment=BOTTOM_ALIGNMENT)
        
        apply_border(sheet, f'{start_col}{start_row}:{end_col}{end_row}', color)
        
    def slot_monitor(sheet:Worksheet, data:dict, timetable_start_col: int, timetable_start_row: int) -> None:
        start_col = col_letter(timetable_start_col + DAY_PARSER[data['dia_semana']]*4 + (4 if (data['posicao'] == 2) else 2))
        end_col = col_letter(timetable_start_col + DAY_PARSER[data['dia_semana']]*4 + (2 if (data['posicao'] == 1) else 4)) 
        start_row = time_to_integer(data['hora_inicio']) + timetable_start_row + 6
        end_row = time_to_integer(data['hora_fim']) + timetable_start_row + 5

        if not conflit(sheet, col_number(start_col), start_row, col_number(end_col), end_row):
            error(f'Não foi possível adicionar a {data["tipo_atividade"]} da disciplina {data["nome_disciplina"]} da turma {data["curso"]}-{get_digit(data["serie"])}{data["turma"]} no horário {data["hora_inicio"]} às {data["hora_fim"]} da {data["dia_semana"]} devido a um conflito com outro disciplina já adicionada')
            return
        duration = time_differece(data['hora_inicio'], data['hora_fim'])
        if duration < 1.5:
            error(f'Não foi possível adicionar a {data["tipo_atividade"]} da disciplina {data["nome_disciplina"]} da turma {data["curso"]}-{get_digit(data["serie"])}{data["turma"]} pois a duração da aula é diferente de 1h30')
            print(f'Horário: {data["hora_inicio"]} às {data["hora_fim"]} da {data["dia_semana"]} | Duração: {float_to_time(duration)}')
        
        color = ATTENDANCE_BORDERS[data['cor']] if data['cor'] < 6 else ATTENDANCE_BORDERS[-1]    
        
        slot_title_cell = merge_cells(sheet, f'{start_col}{start_row}:{end_col}{start_row}')
        cell_styles(slot_title_cell, 'Monitoria', FONT_BOLD_GRAY7)
        
        slot_subtitle_cell = merge_cells(sheet, f'{start_col}{start_row+1}:{end_col}{start_row+1}')
        cell_styles(slot_subtitle_cell, data['docentes'], FONT_BASE7)
        
        slot_middle_cell = merge_cells(sheet, f'{start_col}{start_row+2}:{end_col}{start_row+3}')
        cell_styles(slot_middle_cell, data['nome_disciplina'], FONT_BASE7, alignment=CENTER_ALIGNMENT)
        
        slot_footer_cell = merge_cells(sheet, f'{start_col}{start_row+4}:{end_col}{end_row}')
        cell_styles(slot_footer_cell,  f'{data["hora_inicio"]} às {data["hora_fim"]}', FONT_BASE7, alignment=BOTTOM_ALIGNMENT)
        
        apply_border(sheet, f'{start_col}{start_row}:{end_col}{end_row}', color)

    def full_day_slot(sheet:Worksheet, data:dict, timetable_start_col: int, timetable_start_row: int) -> None:
        start_col = col_letter(timetable_start_col + DAY_PARSER[data['dia_semana']]*4 + 2)
        end_col = col_letter(timetable_start_col + DAY_PARSER[data['dia_semana']]*4 + 4)
        start_row = time_to_integer(data['hora_inicio']) + timetable_start_row + 6
        end_row = time_to_integer(data['hora_fim']) + timetable_start_row + 5
        
        color = ORANGE_FILL
        
        first_row = start_row
        in_class = False
        current_row = start_row
        while current_row <= end_row:
            current_cell = sheet.cell(row=current_row, column=col_number(start_col))
            cell_in_conflit = current_cell.fill.start_color.rgb != WHITE or current_cell.value is not None
            if cell_in_conflit and (not in_class):
                in_class = True
                cell = merge_cells(sheet, f'{start_col}{first_row}:{end_col}{current_row-1}')
                cell.fill = color
                cell_styles(cell, data['nome_disciplina'], FONT_BOLD_GRAY7, color, CENTER_ALIGNMENT)
            elif (not cell_in_conflit) and in_class:
                in_class = False
                first_row = current_row
            current_row += 1
        if (not in_class):
            cell = merge_cells(sheet, f'{start_col}{first_row}:{end_col}{end_row}')
            cell.fill = color
            cell_styles(cell, data['nome_disciplina'], FONT_BOLD_GRAY7, color, CENTER_ALIGNMENT)

    def special_slot(sheet:Worksheet, data:dict, timetable_start_col: int, timetable_start_row: int) -> None:
        start_col = col_letter(timetable_start_col + DAY_PARSER[data['dia_semana']]*4 + 2)
        end_col = col_letter(timetable_start_col + DAY_PARSER[data['dia_semana']]*4 + 4)
        start_row = time_to_integer(data['hora_inicio']) + timetable_start_row + 6
        end_row = time_to_integer(data['hora_fim']) + timetable_start_row + 5
        
        if not conflit(sheet, col_number(start_col), start_row, col_number(end_col), end_row):
            error(f'Não foi possível adicionar a {data['tipo_atividade']} da disciplina {data["nome_disciplina"]} da turma {data['curso']}-{get_digit(data['serie'])}{data['turma']} no horário {data["hora_inicio"]} às {data["hora_fim"]} da {data['dia_semana']} devido a um conflito com outro disciplina já adicionada')
            return

        cell_title = merge_cells(sheet, f'{start_col}{start_row}:{end_col}{end_row-1}')
        cell_styles(cell_title, data['nome_disciplina'], FONT_BOLD_GRAY7, LIGHT_ORANGE_FILL, CENTER_ALIGNMENT)
        cell_time = merge_cells(sheet, f'{start_col}{end_row}:{end_col}{end_row}')
        cell_styles(cell_time, f'{data["hora_inicio"]} às {data["hora_fim"]}', FONT_BOLD_RED7, LIGHT_ORANGE_FILL, CENTER_ALIGNMENT)
  
    
    for idx_course, course_timetable in enumerate(COURSE_TIMETABLES):
        wb = openpyxl.Workbook()
        excel_name = f'GRADE HORÁRIA {course_timetable.upper()} {CURRENT_YEAR}-{CURRENT_SEMESTER}'
        info(f'Criando arquivo {excel_name}')
        
        for sheet_type in SHEET_TYPES:
            sheet_name = f'GRADE {CURRENT_YEAR}-{CURRENT_SEMESTER} {sheet_type}'
            with_attendance = sheet_type == 'COM ATEND.'
            with_professors = sheet_type != 'SEM PROF.'
            
            info(f'Criando {sheet_name}')
            wb.create_sheet(sheet_name)
            sheet = wb[sheet_name]
            
            data_course = DATA[DATA['curso'].isin(COURSE_FILTERS[idx_course])]
            series = sorted(data_course['serie'].unique(), key=lambda x: str(x))
            n_series = len(series)
            n_turmas = data_course['turma'].nunique()
            
            for row in sheet.iter_rows(min_row=1, max_row=timetable_height*n_turmas+2, min_col=1, max_col=timetable_width*n_series):
                for cell in row:
                    cell_styles(cell, font=FONT_BASE7, fill=WHITE_FILL, alignment=CENTER_ALIGNMENT)
            
            for idx_serie, serie in enumerate(series):
                serie_number = get_digit(serie)
                
                timetable_start_col = timetable_width*idx_serie + 1
                timetable_end_col = timetable_width*idx_serie + 21
                
                data_serie = data_course[(DATA['serie'] == serie)].sort_values(['curso', 'turma'])   
                
                turmas = data_serie.groupby(['curso', 'turma'])['nome_disciplina'].nunique().reset_index().rename(columns={'nome_disciplina': 'Quant. nome_disciplinas'})
                turmas['nome_disciplinas'] = data_serie.groupby(['curso', 'turma'])['nome_disciplina'].apply(lambda x: ', '.join(set(x.dropna()))).reset_index()['nome_disciplina']
                turmas = turmas[(turmas['Quant. nome_disciplinas'] > 1) | (~turmas['nome_disciplinas'].isin(CONFIGS['DISCIPLINES_GRUPED_CLASS']))]
                
                if turmas.shape[0] == 0:
                    continue
                title_range = f'{col_letter(timetable_start_col)}1:{col_letter(timetable_end_col)}1'
                cell = merge_cells(sheet, title_range)
                cell_styles(cell, f'Grade Horária\n{course_timetable}\n{CURRENT_YEAR}-{CURRENT_SEMESTER}', FONT_BOLD_RED10)
                sheet.row_dimensions[1].height = TITLE_HEIGHT
                apply_border(sheet, title_range, DEFAULT_BORDER)
            
                LOGO = Image(path.join(abs_path, 'img', 'insper.png'))
                LOGO.width = 3.5 * CM2PIXEL  # Largura em pixels
                LOGO.height = 1.5 * CM2PIXEL  # Altura em pixels
                sheet.add_image(LOGO, title_range.split(':')[0])
                
                for idx_turma, turma_row in turmas.iterrows():
                    data_turma = data_serie[(data_serie['turma'] == turma_row['turma']) & (data_serie['curso'] == turma_row['curso'])]\
                        .sort_values(['nome_disciplina'])\
                        .copy()
                        
                    reserved_day_disciplines = data_turma[data_turma['tipo_atividade'] == 'DIA RESERVADO']['nome_disciplina'].unique()
                    data_turma.loc[data_turma['nome_disciplina'].isin(reserved_day_disciplines), 'cor'] = 6
                    if 'DEVELOPER LIFE' in data_turma['nome_disciplina'].values:
                        data_turma.loc[data_turma['nome_disciplina'].str.contains('DEVELOPER LIFE'), 'cor'] = data_turma[data_turma['nome_disciplina'] == 'DEVELOPER LIFE']['cor'].unique()[0]
                    
                    course_title = TITLE_COURSE[turmas.loc[idx_turma, 'curso']]
                    
                    timetable_start_row = timetable_height*idx_turma + 1 
                    
                    subtitle_range = f'{col_letter(timetable_start_col)}{timetable_start_row+2}:{col_letter(timetable_end_col)}{timetable_start_row+2}'
                    cell = merge_cells(sheet, subtitle_range)
                    cell_styles(cell, f'{serie_number}º Período {"-" if course_title == "" else "- "+course_title+" -"} Turma {turma_row['turma'].replace('Z@', '')}', FONT_BOLD_BLACK8)
                    sheet.row_dimensions[3].height = DATA_HEIGHT
                    sheet.row_dimensions[4].height = SPACING_HEIGHT
                    apply_border(sheet, subtitle_range, RED_BORDER)
                    
                    # Adiciona os dias da semana na primeira linha
                    cell = sheet.cell(timetable_start_row+4, timetable_start_col)
                    cell_styles(cell, 'Horário', FONT_BOLD_WHITE8, RED_FILL)
                    sheet.column_dimensions[col_letter(timetable_start_col)].width = DATA_WIDTH
                    sheet.column_dimensions[col_letter(timetable_start_col+1)].width = SPACING_WIDTH

                    sheet.row_dimensions[timetable_start_row+5].height = SPACING_HEIGHT

                    for i, times in enumerate(zip(CONFIGS['FIXED_START_HOURS'], CONFIGS['FIXED_END_HOURS'])):
                        start_time, end_time = times  
                        start_row = time_to_integer(start_time) + timetable_start_row + 6
                        end_row = time_to_integer(end_time) + timetable_start_row + 5
                        
                        cell = merge_cells(sheet, f'{col_letter(timetable_start_col)}{start_row}:{col_letter(timetable_start_col)}{end_row}')
                        slot = f'{start_time}\nàs\n{end_time}'.replace(':', 'h').replace('00', '')
                        cell_styles(cell, slot, FONT_BOLD_GRAY7, GRAY_FILL)
                    
                    for i, day in enumerate(DAYS):
                        start_col = timetable_start_col + i*4 + 2
                        cell = merge_cells(sheet, f'{col_letter(start_col)}{timetable_start_row+4}:{col_letter(start_col+2)}{timetable_start_row+4}')
                        cell_styles(cell, day, FONT_BOLD_WHITE8, RED_FILL)
                        sheet.column_dimensions[col_letter(start_col)].width = CALENDAR_WIDTH
                        sheet.column_dimensions[col_letter(start_col+1)].width = SPACING_WIDTH
                        sheet.column_dimensions[col_letter(start_col+2)].width = CALENDAR_WIDTH
                        sheet.column_dimensions[col_letter(start_col+3)].width = SPACING_WIDTH
                        
                    sheet.column_dimensions[col_letter(timetable_end_col+1)].width = DIVISION_WIDTH
                    
                    try:
                        for day in DAYS:
                            data_day = data_turma[data_turma['dia_semana'] == f'{day}-FEIRA']
                            data_day_in_class = data_day[data_day['tipo_atividade'] != 'DIA RESERVADO']
                            for i, data_slot in data_day_in_class.iterrows():
                                if data_slot['posicao'] < 0:
                                    continue
                                if data_slot['nome_disciplina'] in CONFIGS['SPECIAL_CLASSES']:
                                    special_slot(sheet, data_slot.to_dict(), timetable_start_col, timetable_start_row, )
                                    
                                elif (data_slot['tipo_atividade'] == 'ATENDIMENTO / PLANTÃO') and with_attendance:
                                    slot_attendance(sheet, data_slot.to_dict(), timetable_start_col, timetable_start_row)
                                    
                                elif (data_slot['tipo_atividade'] in ['MONITORIA', 'MONITORIA NINJA']) and with_attendance:
                                    slot_monitor(sheet, data_slot.to_dict(), timetable_start_col, timetable_start_row)
                                    
                                elif (data_slot['tipo_atividade'] in ['AULA', 'ATIVIDADE EXTRA CURRICULAR']):
                                    slot_class(sheet, data_slot.to_dict(), timetable_start_col, timetable_start_row, with_professors, with_attendance) 
                                                        
                            data_day_reserved = data_day[data_day['tipo_atividade'] == 'DIA RESERVADO']
                            if data_day_reserved.shape[0] > 0:
                                full_day_slot(sheet, data_day_reserved.iloc[0].to_dict(), timetable_start_col, timetable_start_row)
                                if data_day_reserved.shape[0] > 1:
                                    warning(f'Dia reservado com mais de uma vez na {day}-FEIRA para turma {turma_row["curso"]}-{serie_number}{turma_row["turma"]}')
                    except Exception as e:
                        error(f'Erro ao adicionar aulas na {day}-FEIRA para turma {turma_row["curso"]}-{serie_number}{turma_row["turma"]}.\nErro: {e}')
                        error(f'Dados:\n{data_slot}')
                        
                footer_range = f'{col_letter(timetable_start_col)}{timetable_start_row+timetable_height+1}:{col_letter(timetable_end_col)}{timetable_start_row+timetable_height+1}'
                cell = merge_cells(sheet, footer_range)
                cell_styles(cell, 'A grade horária está sujeita a alterações até o início das aulas', FONT_BOLD_BLACK8)
                apply_border(sheet, footer_range, DEFAULT_BORDER)
                
                sheet.row_breaks.append(Break(id=timetable_start_row+timetable_height+1))
                sheet.col_breaks.append(Break(id=timetable_end_col+1))
            
        wb.remove(wb['Sheet'])
        try:
            wb.save(f'{excel_name}.xlsx')
        except PermissionError as e:
            warning(f'Não foi possível salvar o arquivo {excel_name}.xlsx pois o arquivo está aberto. Feche o arquivo e tente novamente.')
            return
        success('Arquivo criado com sucesso!\n')
    success('Automação finalizada! Pressione [ENTER] para sair')
    input()
