
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.pagebreak import Break

from datetime import datetime
import warnings

warnings.filterwarnings('ignore')

from config_grade import load_configs

# Definições de Tamanhos
CM2PIXEL = 37.7952755906

TITLE_HEIGHT = 43.5

DATA_HEIGHT = 18.8
DATA_WIDTH = 7.44

CALENDAR_HEIGHT = 13.5
CALENDAR_WIDTH = 14

SPACING_HEIGHT = 4.5
SPACING_WIDTH = 0.44
DIVISION_WIDTH = 2

# Definições de Cores
DARK_RED = '00C00000'
RED = '00FF0000'
BLACK = '00000000'
WHITE = '00FFFFFF'
DARK_GRAY = '00595959'
GRAY = '00BFBFBF'
LIGHT_BLACK = '00404040'
ORANGE = '00FFC000'
LIGHT_ORANGE = '00FFDD71'

DISCIPLINES_COLORS = [
    'FABF8F',
    'E6B8B7',
    'C4D79B',
    'B1A0C7',
    '8DB4E2',
    'BFBFBF',
    'FFC000'
]

# Definições de Variaveis
DAYS = ['SEGUNDA', 'TERÇA', 'QUARTA', 'QUINTA', 'SEXTA']
DAY_PARSER = {
    'SEGUNDA-FEIRA': 0,
    'TERÇA-FEIRA': 1,
    'QUARTA-FEIRA': 2,
    'QUINTA-FEIRA': 3,
    'SEXTA-FEIRA': 4
}

CURRENT_DATE = datetime.now()
CURRENT_DATE = datetime.strptime('2024-08-05', '%Y-%m-%d')
CURRENT_YEAR = CURRENT_DATE.year
CURRENT_SEMESTER = 1 if CURRENT_DATE.month < 6 else 2
CURRENT_YEAR = 2024
CURRENT_SEMESTER = 2

COURSE_TIMETABLES = [
    'Engenharias',
    'Administração e Ciências Econômicas',
    'Ciência da Computação',
    'Direito',
]
COURSE_FILTERS = [
    ['ENG', 'COMP', 'MEC/MECAT', 'MEC', 'MECAT'],
    ['ADM/ECO', 'ADM', 'ECO'],
    ['CIECOMP'],
    ['DIR'],
]
TITLE_COURSE = {
    'ENG': '',
    'COMP': 'COMP',
    'MEC/MECAT': 'MEC/MECAT',
    'MEC': 'MEC',
    'MECAT': 'MECAT',
    'ADM/ECO': 'ADM/ECO',
    'ADM': 'Administração',
    'ECO': 'Ciências Econômicas',
    'CIECOMP': '',
    'DIR': '',
}
SHEET_TYPES = [
    'SEM PROF.',
    'COM PROF.',
    'COM ATEND.',
]

CONFIGS = load_configs()

TYPE_PRIORITY = {
    'AULA': 3,
    'ATIVIDADE EXTRA CURRICULAR': 3,
    'ATENDIMENTO / PLANTÃO': 2,
    'MONITORIA': 1,
    'MONITORIA NINJA': 0,
    'DIA RESERVADO': -1,
    'BANCA / APRESENTAÇÃO': -2,
}

# Definição de horários fixos para aulas
FIXED_START_HOURS = ['07:30', '09:45', '12:00', '14:15', '16:30', '19:00'] if CONFIGS['NEW_TIMETABLE'] else  ['07:30', '09:45', '12:00', '13:30', '15:45', '18:00'] 
FIXED_END_HOURS = ['09:30', '11:45', '14:00', '16:15', '18:30', '21:00'] if CONFIGS['NEW_TIMETABLE'] else ['09:30', '11:45', '13:15', '15:30', '17:45' ,'20:00']

def col_number(column: str) -> int:
    return column_index_from_string(column)
def col_letter(column: int) -> str:
    return get_column_letter(column)

def parse_range(range:str) -> tuple[int, int, int, int]:
    start, end = range.split(':')
    start_row, start_col = ''.join(filter(str.isdigit, start)), ''.join(filter(str.isalpha, start))
    end_row, end_col = ''.join(filter(str.isdigit, end)), ''.join(filter(str.isalpha, end))
    return int(start_row), col_number(start_col), int(end_row), col_number(end_col)

def apply_border(sheet: Worksheet, range: str, border: Border) -> None:
    start_row, start_col, end_row, end_col = parse_range(range)
    border_list = isinstance(border, list)
    cells = list(sheet.iter_rows(min_row=start_row, min_col=start_col, max_row=end_row, max_col=end_col))
    for i, row in enumerate(cells):
        for cell in row:
            if border_list:
                if i == 0: cell.border = border[0]
                elif i == len(cells)-1: cell.border = border[2]
                else: cell.border = border[1]
            else:
                cell.border = border 
                
def merge_cells(sheet: Worksheet, range:str) -> Cell:
    start_row, start_col, _, _ = parse_range(range)
    sheet.merge_cells(range)
    return sheet.cell(row=start_row, column=start_col)

def cell_styles(cell:Cell, value:str=None, font:Font=None, fill:PatternFill=None, alignment:Alignment=None) -> None:
    if value is not None: cell.value = value
    if font is not None: cell.font = font
    if fill is not None: cell.fill = fill
    if alignment is not None: cell.alignment = alignment
    
    
# Define os estilos das aulas e atendimentos
CLASS_COLORS = [
    PatternFill(start_color=color, end_color=color, fill_type="solid")
    for color in DISCIPLINES_COLORS
]

ATTENDANCE_BORDERS = []
for color in DISCIPLINES_COLORS:
    side = Side(border_style='thick', color=color)
    ATTENDANCE_BORDERS.append([
        Border(side, side, side, None),
        Border(side, side, None, None),
        Border(side, side, None, side),
    ])

# Define os estilos de borda
DEFAULT_BORDER = Border(
    left=Side(border_style='thin', color=BLACK),
    right=Side(border_style='thin', color=BLACK),
    top=Side(border_style='thin', color=BLACK),
    bottom=Side(border_style='thin', color=BLACK)
)
RED_BORDER = Border(
    left=Side(border_style='thin', color=DARK_RED),
    right=Side(border_style='thin', color=DARK_RED),
    top=Side(border_style='thin', color=DARK_RED),
    bottom=Side(border_style='thin', color=DARK_RED)
)

# Define os estilos de preenchimento
WHITE_FILL = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
RED_FILL = PatternFill(start_color=DARK_RED, end_color=DARK_RED, fill_type="solid")
GRAY_FILL = PatternFill(start_color=GRAY, end_color=GRAY, fill_type="solid")
ORANGE_FILL = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
LIGHT_ORANGE_FILL = PatternFill(start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type="solid")

# Define os estilos de fonte
FONT_BOLD_RED10 = Font(color=DARK_RED, bold=True, size=10, name='Arial')
FONT_BOLD_BLACK8 = Font(color=DARK_GRAY, bold=True, size=8, name='Arial')
FONT_BOLD_WHITE8 = Font(color=WHITE, bold=True, size=8, name='Arial')
FONT_BOLD_GRAY7 = Font(color=LIGHT_BLACK, bold=True, size=7, name='Arial')
FONT_BOLD_RED7 = Font(color=RED, bold=True, size=7, name='Arial')
FONT_BASE7 = Font(color=BLACK, bold=False, size=7, name='Arial')

# Define os estilos de alinhamento
CENTER_ALIGNMENT = Alignment(wrap_text=True, horizontal="center", vertical="center")
TOP_ALIGNMENT = Alignment(wrap_text=True, horizontal="center", vertical="top")
BOTTOM_ALIGNMENT = Alignment(wrap_text=True, horizontal="center", vertical="bottom")